"""Generación de reportes Excel (consumidos por n8n para envío por correo)."""
from io import BytesIO
from datetime import date
from typing import Optional, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from . import models
from .constants import EMBALAJE_ITEMS_F006, TIPOS_PRUEBA_F006, TIPOS_MATERIAL_F006

_EMB_LABEL = dict(EMBALAJE_ITEMS_F006)
_PRUEBA_LABEL = dict(TIPOS_PRUEBA_F006)
_MATERIAL_LABEL = dict(TIPOS_MATERIAL_F006)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)


def _style_header(ws, row_idx: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_f006_workbook(registros: List[models.F006Registro], fecha: date, turno: Optional[int]) -> BytesIO:
    wb = Workbook()

    # Hoja 1: Filtración
    ws = wb.active
    ws.title = "Filtración"
    ws["A1"] = "F-006 · Ruta control proceso vasos — Filtración"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Fecha: {fecha}" + (f"  ·  Turno: {turno}" if turno else "")

    headers = [
        "Fecha", "Turno", "Máquina", "Referencia", "Hora montaje", "Tipo de prueba",
        "Material", "Cant. muestra", "Hora lectura", "Cumple (no filtra)",
        "No cumple (filtra)", "Estado", "Comentario",
    ]
    hrow = 4
    ws.append([])  # fila 3 vacía
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    _style_header(ws, hrow, len(headers))

    r = hrow + 1
    for reg in registros:
        maquina = reg.maquina.nombre if reg.maquina else ""
        referencia = reg.referencia.codigo if reg.referencia else ""
        for f in reg.filtraciones:
            ws.cell(row=r, column=1, value=str(reg.fecha))
            ws.cell(row=r, column=2, value=reg.turno)
            ws.cell(row=r, column=3, value=maquina)
            ws.cell(row=r, column=4, value=referencia)
            ws.cell(row=r, column=5, value=f.hora_montaje.strftime("%H:%M") if f.hora_montaje else "")
            ws.cell(row=r, column=6, value=_PRUEBA_LABEL.get(f.tipo_prueba, f.tipo_prueba))
            ws.cell(row=r, column=7, value=_MATERIAL_LABEL.get(f.tipo_material, f.tipo_material))
            ws.cell(row=r, column=8, value=f.cantidad_muestra)
            ws.cell(row=r, column=9, value=f.hora_lectura.strftime("%H:%M") if f.hora_lectura else "")
            ws.cell(row=r, column=10, value=f.cantidad_cumple)
            ws.cell(row=r, column=11, value=f.cantidad_nocumple)
            ws.cell(row=r, column=12, value=f.estado)
            ws.cell(row=r, column=13, value=f.comentario or "")
            r += 1

    for col, width in zip("ABCDEFGHIJKLM", [12, 7, 16, 14, 12, 14, 12, 12, 12, 16, 16, 12, 40]):
        ws.column_dimensions[col].width = width

    # Hoja 2: Embalaje
    ws2 = wb.create_sheet("Embalaje")
    ws2["A1"] = "F-006 · Revisión de embalaje"
    ws2["A1"].font = _TITLE_FONT
    emb_headers = ["Máquina", "Referencia", "Turno"] + [lbl for _, lbl in EMBALAJE_ITEMS_F006]
    for i, h in enumerate(emb_headers, start=1):
        ws2.cell(row=3, column=i, value=h)
    _style_header(ws2, 3, len(emb_headers))
    r = 4
    for reg in registros:
        row = [
            reg.maquina.nombre if reg.maquina else "",
            reg.referencia.codigo if reg.referencia else "",
            reg.turno,
        ]
        resultados = {e.item: e.resultado for e in reg.embalaje}
        for key, _ in EMBALAJE_ITEMS_F006:
            row.append(resultados.get(key, ""))
        for i, val in enumerate(row, start=1):
            ws2.cell(row=r, column=i, value=val)
        r += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def build_f015_workbook(mediciones: List[models.F015Medicion], fecha: date) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cloro y PH"
    ws["A1"] = "F-015 · Medición de cloro y PH del agua"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Fecha: {fecha}  ·  Rangos: PH 6.5–8.5  ·  Cloro 0.3–2"

    headers = ["Fecha y hora", "Punto de medición", "PH", "PH en rango", "Cloro",
               "Cloro en rango", "Responsable", "Comentario"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header(ws, 4, len(headers))

    r = 5
    fuera_fill = PatternFill("solid", fgColor="FFC7CE")
    for m in mediciones:
        ws.cell(row=r, column=1, value=m.fecha_hora.strftime("%Y-%m-%d %H:%M") if m.fecha_hora else "")
        ws.cell(row=r, column=2, value=m.punto.nombre if m.punto else "")
        c_ph = ws.cell(row=r, column=3, value=m.ph)
        ws.cell(row=r, column=4, value="Sí" if m.ph_en_rango else "NO")
        c_cl = ws.cell(row=r, column=5, value=m.cloro)
        ws.cell(row=r, column=6, value="Sí" if m.cloro_en_rango else "NO")
        ws.cell(row=r, column=7, value=m.responsable.nombre if m.responsable else "")
        ws.cell(row=r, column=8, value=m.comentario or "")
        if not m.ph_en_rango:
            c_ph.fill = fuera_fill
        if not m.cloro_en_rango:
            c_cl.fill = fuera_fill
        r += 1

    for col, width in zip("ABCDEFGH", [20, 26, 8, 12, 8, 12, 20, 40]):
        ws.column_dimensions[col].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
